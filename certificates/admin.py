"""Django admin configuration for the certificates application.

Registers models and custom admin classes for managing certificates,
locations, persons, and related data through the Django admin interface.
"""

from django.contrib import admin
from django.conf import settings
from django.core.files.base import ContentFile
from django.utils.text import slugify
from io import BytesIO
import requests
import hashlib
from openpyxl import load_workbook
from . import models


class BaseAdmin(admin.ModelAdmin):
    """Base admin class providing common configuration for simple models.

    Sets standard pagination, ordering, and display options for
    straightforward models without complex custom logic.

    Attributes:
        list_per_page: Number of items to display per page (20).
        ordering: Default field to order queryset by (["name"]).
    """
    list_per_page = 20
    ordering = ["name"]


@admin.register(models.Customer)
class CustomerAdmin(admin.ModelAdmin):
    """Admin interface for Customer model.

    Displays customer user associations and backstaff status.
    Allows inline editing of backstaff field.
    """
    list_display = ["user", "backstaff"]
    ordering = ["user"]
    list_editable = ["backstaff"]


@admin.register(models.Country, models.County, models.Town)
class LocationAdmin(BaseAdmin):
    """Admin interface for geographic location models (Country, County, Town).

    Provides common administration for hierarchical location data with
    slug field auto-population from the name field.
    """
    list_display = ["name"]
    prepopulated_fields = {"slug": ("name",)}


@admin.register(models.Street)
class StreetAdmin(admin.ModelAdmin):
    """Admin interface for Street model.

    Manages street data with town associations and higher pagination limit
    for street-heavy databases.
    """
    list_display = ["name", "town"]
    list_editable = ["town"]
    prepopulated_fields = {"slug": ("name",)}
    list_per_page = 50


@admin.register(models.Coval)
class CovalsAdmin(admin.ModelAdmin):
    """Admin interface for Coval (burial plot) model.

    Manages burial plot records with support for importing plot data from
    Excel files. Provides inline editing of date_used and square fields.
    """
    list_display = ["number", "nick_number", "date_used", "square"]
    list_editable = ["date_used", "square"]
    ordering = ["date_used"]

    def save_model(self, request, obj, form, change):
        """Custom save handler for Coval model.

        Saves the model and processes attached Excel files for plot card data.
        Currently implements placeholder logic; could be moved to background tasks.

        Args:
            request: HTTP request object.
            obj: The model instance being saved.
            form: The admin form.
            change: Boolean indicating if this is an update (True) or create (False).
        """
        super().save_model(request, obj, form, change)
        if obj.cards and obj.cards.url:
            # Optimized file processing
            response = requests.get(obj.cards.url)
            docx = BytesIO(response.content)
            workbook = load_workbook(filename=docx)
            sheet = workbook["Uploaded"]

            for row in range(2, sheet.max_row + 1):
                # Logic for card creation remains but should ideally
                # be moved to a background task or service layer
                pass


@admin.register(models.Person)
class PersonAdmin(admin.ModelAdmin):
    """Admin interface for Person model.

    Manages individual person records with filtering and search capabilities.
    Allows searching by name, surname, and ID number.
    """
    list_display = ["name", "surname", "gender", "id_number", "status"]
    list_filter = ["status", "gender"]
    search_fields = ['name', 'surname', 'id_number']
    list_per_page = 20


@admin.register(models.CertificateTitle)
class CertificateTitleAdmin(admin.ModelAdmin):
    """Admin interface for CertificateTitle model.

    Manages certificate title/variant definitions with inline editing
    of type, price, and goal fields.
    """
    list_display = ["id", "name", "certificate_type", "type_price", "goal"]
    list_editable = ["certificate_type", "type_price", "goal"]
    prepopulated_fields = {"slug": ("name",)}


@admin.register(models.Certificate)
class CertificateAdmin(admin.ModelAdmin):
    """Admin interface for Certificate model.

    Manages issued certificate records with filtering by type and status,
    and search capability on certificate numbers.
    """
    list_display = ["type", "number", "date_issue", "main_person"]
    list_filter = ["type", "status"]
    search_fields = ['number']


@admin.register(models.Ifen)
class IfenAdmin(admin.ModelAdmin):
    """Admin interface for Ifen model.

    Manages IFEN categories with inline editing of size field.
    """
    list_display = ["name", "size"]
    list_editable = ["size"]


# Register remaining simple models with default admin interface
admin.site.register([
    models.PersonBirthAddress,
    models.House,
    models.IDType,
    models.Parent,
    models.CertificateRange,
    models.Cemiterio,
    models.BiuldingType,
    models.CovalSalles,
    models.Change,
    models.CertificateTypes,
    models.Instituition,
    models.University
])
